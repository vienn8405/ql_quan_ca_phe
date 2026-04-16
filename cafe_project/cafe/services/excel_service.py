"""
Service xử lý import/export Excel cho hệ thống quản lý quán cà phê.
"""
import io
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse


# ======================== IMPORT NGUYÊN VẬT LIỆU ========================

MATERIAL_HEADERS = ['name', 'unit', 'quantity', 'min_quantity']
VALID_UNITS = ['g', 'ml', 'item']


def generate_material_template():
    """Tạo file Excel mẫu cho import nguyên vật liệu."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Nguyên vật liệu"

    # Header
    headers = ['name', 'unit', 'quantity', 'min_quantity']
    header_labels = ['Tên nguyên liệu', 'Đơn vị (g/ml/item)', 'Số lượng', 'Mức tối thiểu']
    for col, (key, label) in enumerate(zip(headers, header_labels), 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = cell.font.copy(bold=True)

    # Dữ liệu mẫu
    sample_data = [
        ['Cà phê rang xay', 'g', 5000, 500],
        ['Sữa tươi', 'ml', 10000, 2000],
        ['Ly giấy', 'item', 200, 50],
    ]
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Điều chỉnh chiều rộng cột
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18

    # Ghi vào buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mau_import_nguyen_vat_lieu.xlsx"'
    return response


def import_materials_from_excel(file, branch=None):
    """
    Đọc file Excel và trả về kết quả import nguyên vật liệu.

    Logic cập nhật:
    - Nếu nguyên liệu chưa tồn tại (theo tên) → tạo mới
    - Nếu đã tồn tại → GHI ĐÈ quantity, unit, min_quantity theo file

    Returns:
        dict: {
            'created_count': int,
            'updated_count': int,
            'error_messages': list[str],
            'success': bool
        }
    """
    from cafe.models import RawMaterial, BranchMaterialStock, StockLog

    result = {
        'created_count': 0,
        'updated_count': 0,
        'error_messages': [],
        'success': True,
    }

    # Đọc file Excel
    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        result['success'] = False
        result['error_messages'].append(f'Không thể đọc file Excel: {str(e)}')
        return result

    ws = wb.active

    # Kiểm tra header
    header_row = [str(cell.value or '').strip().lower() for cell in ws[1]]

    # Map header linh hoạt: hỗ trợ cả tiếng Việt và tiếng Anh
    header_map = {}
    name_aliases = ['name', 'tên nguyên liệu', 'tên', 'ten nguyen lieu', 'ten', 'tên nguyên liệu (*)', 'nguyên liệu', 'nguyen lieu']
    unit_aliases = ['unit', 'đơn vị', 'don vi', 'đơn vị (g/ml/item)', 'don vi (g/ml/item)', 'dv']
    quantity_aliases = ['quantity', 'số lượng', 'so luong', 'sl', 'số lượng tồn', 'so luong ton']
    min_qty_aliases = ['min_quantity', 'mức tối thiểu', 'muc toi thieu', 'min', 'tối thiểu', 'toi thieu']

    for idx, header in enumerate(header_row):
        h = header.lower().strip()
        if h in name_aliases:
            header_map['name'] = idx
        elif h in unit_aliases:
            header_map['unit'] = idx
        elif h in quantity_aliases:
            header_map['quantity'] = idx
        elif h in min_qty_aliases:
            header_map['min_quantity'] = idx

    if 'name' not in header_map:
        result['success'] = False
        result['error_messages'].append(
            'Không tìm thấy cột "Tên nguyên liệu" trong file. '
            'Header cần có: Tên nguyên liệu, Đơn vị, Số lượng, Mức tối thiểu'
        )
        return result

    # Đọc từng dòng dữ liệu (bỏ header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Bỏ qua dòng trống
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue

        # Đọc name
        name_idx = header_map['name']
        name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] else ''
        if not name:
            result['error_messages'].append(f'Dòng {row_idx}: Thiếu tên nguyên liệu → bỏ qua.')
            continue

        # Đọc unit
        unit = 'g'  # mặc định
        if 'unit' in header_map:
            unit_idx = header_map['unit']
            if unit_idx < len(row) and row[unit_idx]:
                raw_unit = str(row[unit_idx]).strip().lower()
                # Map tiếng Việt
                unit_map = {'g': 'g', 'gram': 'g', 'ml': 'ml', 'item': 'item', 'cái': 'item', 'cai': 'item'}
                unit = unit_map.get(raw_unit, raw_unit)
                if unit not in VALID_UNITS:
                    result['error_messages'].append(
                        f'Dòng {row_idx}: Đơn vị "{raw_unit}" không hợp lệ (chỉ g/ml/item). Dùng mặc định "g".'
                    )
                    unit = 'g'

        # Đọc quantity
        quantity = 0
        if 'quantity' in header_map:
            qty_idx = header_map['quantity']
            if qty_idx < len(row) and row[qty_idx] is not None:
                try:
                    quantity = float(row[qty_idx])
                    if quantity < 0:
                        result['error_messages'].append(
                            f'Dòng {row_idx}: Số lượng âm ({quantity}) → đặt về 0.'
                        )
                        quantity = 0
                except (ValueError, TypeError):
                    result['error_messages'].append(
                        f'Dòng {row_idx}: Số lượng "{row[qty_idx]}" không hợp lệ → đặt về 0.'
                    )
                    quantity = 0

        # Đọc min_quantity
        min_quantity = 0
        if 'min_quantity' in header_map:
            mq_idx = header_map['min_quantity']
            if mq_idx < len(row) and row[mq_idx] is not None:
                try:
                    min_quantity = float(row[mq_idx])
                    if min_quantity < 0:
                        min_quantity = 0
                except (ValueError, TypeError):
                    result['error_messages'].append(
                        f'Dòng {row_idx}: Mức tối thiểu "{row[mq_idx]}" không hợp lệ → đặt về 0.'
                    )
                    min_quantity = 0

        # Tạo mới hoặc cập nhật (theo tên, case-insensitive)
        try:
            existing = RawMaterial.objects.filter(name__iexact=name).first()
            if existing:
                material = existing
                material.unit = unit
                if branch is None:
                    material.quantity = quantity
                    material.min_quantity = min_quantity
                material.save()
                is_created = False
            else:
                material = RawMaterial.objects.create(
                    name=name,
                    unit=unit,
                    quantity=quantity,
                    min_quantity=min_quantity,
                )
                is_created = True

            if branch is not None:
                branch_stock, _ = BranchMaterialStock.objects.get_or_create(
                    material=material,
                    branch=branch,
                    defaults={'quantity': quantity, 'min_quantity': min_quantity}
                )
                branch_stock.quantity = quantity
                branch_stock.min_quantity = min_quantity
                branch_stock.save()

                material.quantity = sum(stock.quantity for stock in material.branch_stocks.all())
                material.min_quantity = max((stock.min_quantity for stock in material.branch_stocks.all()), default=0)
                material.save(update_fields=['quantity', 'min_quantity', 'updated_at'])

                StockLog.objects.create(
                    stock_type='raw',
                    action='import',
                    material=material,
                    branch=branch,
                    quantity=quantity,
                    note=f'Import Excel cho chi nhánh {branch.name}'
                )

            if is_created:
                result['created_count'] += 1
            else:
                result['updated_count'] += 1
        except Exception as e:
            result['error_messages'].append(f'Dòng {row_idx}: Lỗi lưu "{name}" → {str(e)}')

    return result


# ======================== EXPORT ĐƠN HÀNG ========================

def export_orders_excel(queryset):
    """
    Xuất danh sách đơn hàng ra file Excel (.xlsx).

    Args:
        queryset: QuerySet của Order (có thể đã filter)

    Returns:
        HttpResponse chứa file .xlsx
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "DonHang"

    # ===== HEADER =====
    headers = [
        'Mã đơn',
        'Tên khách hàng',
        'Số điện thoại',
        'Loại đơn',
        'Chi nhánh',
        'Trạng thái',
        'Trạng thái giao',
        'Thanh toán',
        'Phí giao (đ)',
        'Giảm giá (đ)',
        'Tổng tiền (đ)',
        'Ghi chú',
        'Ngày tạo',
    ]

    # Style cho header
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='5B3A1A', end_color='5B3A1A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ===== DATA =====
    # Map display values
    order_type_map = dict(getattr(queryset.model, 'ORDER_TYPE_CHOICES', []))
    status_map = dict(getattr(queryset.model, 'STATUS_CHOICES', []))
    delivery_status_map = dict(getattr(queryset.model, 'DELIVERY_STATUS_CHOICES', []))
    payment_map = dict(getattr(queryset.model, 'PAYMENT_CHOICES', []))

    data_alignment = Alignment(vertical='center', wrap_text=True)
    money_format = '#,##0'
    date_format = 'DD/MM/YYYY HH:MM'

    for row_idx, order in enumerate(queryset.select_related('branch', 'voucher'), start=2):
        row_data = [
            order.id,
            order.customer_name or 'Khách vãng lai',
            order.customer_phone or '',
            order_type_map.get(order.order_type, order.order_type),
            order.branch.name if order.branch else 'Chưa gán',
            status_map.get(order.status, order.status),
            delivery_status_map.get(order.delivery_status, order.delivery_status),
            payment_map.get(order.payment_method, order.payment_method),
            order.delivery_fee,
            order.discount_amount,
            order.total_price,
            order.note or '',
            order.created_at.replace(tzinfo=None) if order.created_at else None,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border

            # Format tiền
            if col_idx in (9, 10, 11):  # phí giao, giảm giá, tổng tiền
                cell.number_format = money_format
            # Format ngày
            elif col_idx == 13:
                cell.number_format = date_format

    # ===== AUTO WIDTH =====
    col_widths = [10, 22, 16, 14, 22, 16, 16, 14, 14, 14, 16, 25, 20]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # ===== RESPONSE =====
    timestamp = datetime.now().strftime('%d%m%Y_%H%M')
    filename = f'danh_sach_don_hang_{timestamp}.xlsx'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ======================== EXPORT BÁO CÁO DOANH THU ========================

def export_revenue_excel(data):
    """
    Xuất báo cáo doanh thu ra file Excel (.xlsx).

    Args:
        data: dict chứa các key từ revenue_report view / _get_revenue_data:
            - date_from, date_to
            - total_revenue (net_order_revenue), total_orders, avg_daily
            - gross_item_revenue, discount_total, delivery_fee_total
            - labels, revenues, order_counts  (doanh thu theo ngày)
            - top_products (theo số lượng)
            - top_products_by_revenue (theo doanh thu)
            - type_labels, type_counts, type_revenues  (theo loại đơn)
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "BaoCaoDoanhThu"

    # Styles
    title_font = Font(bold=True, size=16, color='5B3A1A')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='5B3A1A', end_color='5B3A1A', fill_type='solid')
    label_font = Font(bold=True, size=11, color='5B3A1A')
    money_format = '#,##0'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center')

    row = 1

    # ===== TITLE =====
    ws.merge_cells('A1:G1')
    cell = ws.cell(row=1, column=1, value='BAO CAO DOANH THU - CAFEFLOW')
    cell.font = title_font
    cell.alignment = center_align
    row = 2

    ws.merge_cells('A2:G2')
    ts = datetime.now().strftime('%d/%m/%Y %H:%M')
    cell = ws.cell(row=2, column=1, value=f'Xuat bao cao: {ts}')
    cell.font = Font(size=10, italic=True, color='808080')
    cell.alignment = center_align
    row = 3

    date_from = data.get('date_from', '')
    date_to = data.get('date_to', '')
    ws.merge_cells('A3:G3')
    cell = ws.cell(row=3, column=1, value=f'Thoi gian: {date_from} den {date_to}')
    cell.font = Font(size=10, color='808080')
    cell.alignment = center_align
    row = 5

    # ===== SUMMARY - 4 CHỈ SỐ CHÍNH =====
    ws.cell(row=row, column=1, value='TOM TAT DOANH THU').font = Font(bold=True, size=12, color='5B3A1A')
    row += 1

    summary_items = [
        ('Doanh thu thuc thu (net):', data.get('total_revenue', 0)),
        ('Tong doanh thu mon (gross):', data.get('gross_item_revenue', 0)),
        ('Tong giam gia voucher:', data.get('discount_total', 0)),
        ('Tong phi giao hang:', data.get('delivery_fee_total', 0)),
        ('Tong so don hoan thanh:', data.get('total_orders', 0)),
        ('TB doanh thu / ngay:', data.get('avg_daily', 0)),
    ]
    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label).font = label_font
        c = ws.cell(row=row, column=2, value=value)
        c.font = Font(bold=True, size=12)
        if isinstance(value, (int, float)) and label != 'Tong so don hoan thanh:':
            c.number_format = money_format
        row += 1

    row += 1

    # ===== TOP SẢN PHẨM - 2 BẢNG =====
    top_products = data.get('top_products', [])
    top_by_rev = data.get('top_products_by_revenue', [])

    if top_products or top_by_rev:
        ws.cell(row=row, column=1, value='TOP SAN PHAM').font = Font(bold=True, size=12, color='5B3A1A')
        row += 1

        # Bảng 1: Theo số lượng
        ws.cell(row=row, column=1, value='Theo so luong ban').font = Font(bold=True, size=11, color='5B3A1A')
        row += 1
        prod_headers = ['#', 'San pham', 'So luong', 'Doanh thu mon (d)']
        for col, h in enumerate(prod_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border
        row += 1

        for idx, p in enumerate(top_products, 1):
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=p.get('product__name', '')).border = thin_border
            ws.cell(row=row, column=3, value=p.get('total_qty', 0)).border = thin_border
            c = ws.cell(row=row, column=4, value=p.get('item_revenue', 0))
            c.number_format = money_format
            c.border = thin_border
            row += 1

        row += 1

        # Bảng 2: Theo doanh thu
        ws.cell(row=row, column=1, value='Theo doanh thu').font = Font(bold=True, size=11, color='5B3A1A')
        row += 1
        for col, h in enumerate(prod_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border
        row += 1

        for idx, p in enumerate(top_by_rev, 1):
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=p.get('product__name', '')).border = thin_border
            ws.cell(row=row, column=3, value=p.get('total_qty', 0)).border = thin_border
            c = ws.cell(row=row, column=4, value=p.get('item_revenue', 0))
            c.number_format = money_format
            c.border = thin_border
            row += 1

        row += 1

    # ===== THEO LOẠI ĐƠN =====
    type_labels = data.get('type_labels', [])
    type_counts = data.get('type_counts', [])
    type_revenues = data.get('type_revenues', [])

    if type_labels:
        ws.cell(row=row, column=1, value='DOANH THU THEO LOAI DON').font = Font(bold=True, size=12, color='5B3A1A')
        row += 1

        type_headers = ['Loai don', 'So don', 'Doanh thu thuc thu (d)']
        for col, h in enumerate(type_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border
        row += 1

        for i, t_label in enumerate(type_labels):
            ws.cell(row=row, column=1, value=t_label).border = thin_border
            ws.cell(row=row, column=2, value=type_counts[i] if i < len(type_counts) else 0).border = thin_border
            c = ws.cell(row=row, column=3, value=type_revenues[i] if i < len(type_revenues) else 0)
            c.number_format = money_format
            c.border = thin_border
            row += 1

    # ===== AUTO WIDTH =====
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16

    # Freeze row 1
    ws.freeze_panes = 'A5'

    # ===== RESPONSE =====
    timestamp = datetime.now().strftime('%d%m%Y_%H%M')
    filename = f'bao_cao_doanh_thu_{timestamp}.xlsx'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
